"""
base_excel_builder.py
모든 설계법 Excel Builder가 공통으로 사용하는 기반 클래스.
- 셀 스타일, 헤더, 섹션 1~5 (단면제원, 재료상수, 강도감소계수, 필요/사용철근량) 공통 구현
- 서브클래스에서 반드시 구현해야 하는 메서드 선언
"""
from abc import ABC, abstractmethod
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from string import ascii_uppercase


class BaseExcelBuilder(ABC):
    """
    USD 계열 설계법 Excel 보고서의 공통 Base 클래스.
    공통 포맷/섹션은 여기서 구현하고,
    철근비 검토(섹션6), 전단(섹션10/11), 균열(섹션11/12)은 서브클래스에서 구현한다.
    """

    def __init__(self, analyzer):
        self.analyzer = analyzer
        self.std = analyzer.standard

    def add_to_workbook(self, wb, sheet_name):
        """워크북에 보고서 시트를 추가한다."""
        wsout = wb.create_sheet(title=sheet_name)
        self._setup_sheet(wsout)
        self._write_header(wsout)
        self._write_section1(wsout)
        self._write_section2(wsout)
        self._write_section3(wsout)
        self._write_section4(wsout)
        self._write_section5(wsout)
        row_offset = self._write_section6(wsout)  # 설계법별 구현
        self._write_flexure_strength(wsout, row_offset)
        self._write_shear(wsout, row_offset)
        self._write_crack(wsout, row_offset)

    # ── 공통 초기화 ───────────────────────────────────────────────────────────

    def _setup_sheet(self, ws):
        alpalist = list(ascii_uppercase)
        for i in range(1, 100):
            ws.row_dimensions[i].height = 15
        for col in alpalist:
            ws.column_dimensions[col].width = 3.0
        font_format = Font(size=9, name='굴림체')
        for rows in ws["A1":"Z100"]:
            for cell in rows:
                cell.font = font_format

    def _write_header(self, ws):
        """헤더: 보고서 제목"""
        ws.merge_cells('B1:M1')
        ws['B1'].value = f"[{self.std.name}] - RC 단면 검토 보고서"
        ws['B1'].font = Font(size=11, bold=True, name='굴림체')
        ws['B1'].alignment = Alignment(horizontal='left')

    def _write_section1(self, ws):
        """1) 단면제원 및 설계가정"""
        ana = self.analyzer
        label_f = "Øc" if ana.method == "LSD" else "Øf"
        label_v = "Øs" if ana.method == "LSD" else "Øv"
        phi_f_val = ana.phi_c if ana.method == "LSD" else ana.pi_f
        phi_v_val = ana.phi_s if ana.method == "LSD" else ana.pi_v

        ws['B2'].value = '1) 단면제원 및 설계가정'
        ws['C3'].value = f"fck = {ana.f_ck} MPa, fy = {ana.f_y} MPa, {label_f} = {phi_f_val:.2f}, {label_v} = {phi_v_val:.2f}, Es = {ana.E_s} MPa"

        for r in range(4, 6):
            for c in range(3, 24):
                ws.cell(r, c).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(r, c).border = Border(
                    left=Side(border_style='thin'), right=Side(border_style='thin'),
                    top=Side(border_style='thin'), bottom=Side(border_style='thin')
                )
        for r in range(4, 6):
            for c in [3, 6, 9, 12]:
                ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+2)
        for r in range(4, 6):
            for c in [15, 19, 23]:
                ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+3)
        for c in range(3, 24):
            ws.cell(4, c).fill = PatternFill(fill_type='solid', fgColor='0FFFF0')

        ws['C4'].value = 'B(mm)';   ws['F4'].value = 'H(mm)'; ws['I4'].value = 'd(mm)'
        ws['L4'].value = '피복(mm)'; ws['O4'].value = 'Mu(N.mm)'
        ws['S4'].value = 'Vu(N)';   ws['W4'].value = 'Ms(N.mm)'
        ws['C5'].value = ana.beam_b
        ws['F5'].value = ana.beam_h
        ws['I5'].value = f"{ana.d_eff:.1f}"
        ws['L5'].value = f"{ana.d_c:.1f}"
        ws['O5'].value = ana.Mu_nm
        ws['S5'].value = ana.Vu_n
        ws['W5'].value = ana.Ms_nm

    def _write_section2(self, ws):
        """2) 콘크리트 재료상수"""
        ana = self.analyzer
        ws['B7'].value = '2) 콘크리트 재료상수'
        ws['C8'].value = 'β1    : 등가 사각형 응력 블록의 깊이계수'
        ws['P8'].value = '='
        ws['Q8'].value = ana.beta_1

    def _write_section3(self, ws):
        """3) 강도감소계수(Ø) 산정"""
        ana = self.analyzer
        ws['B10'].value = '3) 강도감소계수(Ø) 산정'
        if ana.method == "LSD":
            ws['C11'].value = f"T = As x fy = {ana.as_use:.3f} x {ana.f_yd:.1f} = {ana.tension_force:.1f} N"
            ws['C12'].value = f"C = alpha_cc x phi_c x fck x alpha x b = {ana.alpha_cc:.2f} x {ana.phi_c:.2f} x {ana.f_ck} x {ana.alpha_fac:.2f} x {ana.beam_b} = {ana.compression_force:.1f} x c"
        else:
            ws['C11'].value = f"T = As x fy = {ana.as_use:.3f} x {ana.f_y:.3f} = {ana.tension_force:.1f} N"
            ws['C12'].value = f"C = 0.85 x fck x a x b = 0.85 x {ana.f_ck:.3f} x a x {ana.beam_b:.3f} = {ana.compression_force:.1f} x a"
        ws['C13'].value = f"T = C 이므로, a = {ana.a:.3f} mm, c = {ana.a:.3f} / β1 = {ana.a:.3f} / {ana.beta_1:.3f} = {ana.c:.3f} mm"
        ws['C14'].value = f"εy = fy / Es = {ana.f_y} / {ana.E_s} = {ana.epsilon_y:.5f}"
        ws['C15'].value = f"εt = 0.00300 x (dt - c) / c = 0.00300 x ({ana.d_eff:.3f} - {ana.c:.2f}) / {ana.c:.2f} = {ana.epsilon_t:.5f}"
        compare = "≥" if ana.epsilon_t >= 0.005 else "<"
        ws['C16'].value = f"εt {compare} 0.0050 이므로 {ana.epsilon_t_result}이며, Ø = {ana.pi_f_r:.2f} 를 적용한다"

    def _write_section4(self, ws):
        """4) 필요철근량 산정"""
        ana = self.analyzer
        ws['B18'].value = '4) 필요철근량 산정'
        ws['C19'].value = 'Mu / Øf = As x fy  x (d - a / 2)              ----------------   ①'
        ws['C20'].value = ' a = As x fy  / ( 0.85 x fck x b)             ----------------   ②'
        ws['C21'].value = ' 식②를 식①에 대입하여 이차방정식으로 As를 구한다'
        ws['E22'].value = ' fy²                                Mu'
        ws['C23'].value = f" ────────── As² - fy x d x As + ───  = 0 ,   Asreq = {ana.as_req:.3f} mm²"
        ws['C24'].value = ' 2 x 0.85 x fck x b                         Øf '

    def _write_section5(self, ws):
        """5) 사용철근량"""
        ana = self.analyzer
        usage = ana.as_use / ana.as_req if ana.as_req > 0 else 9.99
        ws['B26'].value = f"5) 사용철근량 : Asuse = {ana.as_use:.1f} mm², 철근도심 : dc_avg = {ana.d_c:.1f}mm,  [ 사용율 = {usage:.3f} ]"
        ws['F27'].value = f"1단 : {ana.rebar_id} {ana.as_dia1} - {ana.as_num1} EA (= {ana.as_use1:.1f} mm², dc1 = {ana.dc_1:.1f} mm)"
        ws['F28'].value = f"2단 : {ana.rebar_id} {ana.as_dia2} - {ana.as_num2} EA (= {ana.as_use2:.1f} mm², dc2 = {ana.dc_2:.1f} mm)"
        ws['F29'].value = f"3단 : {ana.rebar_id} {ana.as_dia3} - {ana.as_num3} EA (= {ana.as_use3:.1f} mm², dc3 = {ana.dc_3:.1f} mm)"

    # ── 공통 휨강도/전단/균열 섹션 ──────────────────────────────────────────────

    def _write_flexure_strength(self, ws, row_offset):
        """설계 휨강도 산정 섹션 (LSD는 섹션8, USD계열은 섹션7)"""
        ana = self.analyzer
        label_f = "Øc" if ana.method == "LSD" else "Øf"
        phi_f_val = ana.phi_c if ana.method == "LSD" else ana.pi_f
        # LSD spacing check 추가 오프셋
        spacing_offset = 0
        if ana.method == "LSD":
            base_s7 = 39 + row_offset
            ws[f'B{base_s7}'].value = "6-1) 철근 간격 검토"
            ws[f'C{base_s7+1}'].value = f"s_max = min( 2h, 250 ) = {ana.s_detailing_max:.0f} mm"
            ws[f'C{base_s7+2}'].value = f"s = b / n = {ana.beam_b:.1f} / {ana.as_num1} = {ana.s_use:.1f} mm"
            compare_sign_s = "\u2265" if ana.s_detailing_max >= ana.s_use else "<"
            ws[f'C{base_s7+3}'].value = f"s_max {compare_sign_s} s  \u2234 {ana.s_detailing_ok}"
            spacing_offset = 6

        base_s8 = 39 + row_offset + spacing_offset
        num_flex = 7 if ana.method == "LSD" else 6 # If spacing is a sub-item, LSD might want 7 here, but let's follow user 6, 7, 8
        if ana.method == "LSD":
            num_flex = 7 # LSD keeps 1-7 sequential
        else:
            num_flex = 6 # USD follows user 6
        ws[f'B{base_s8}'].value = f"{num_flex}) 설계 휨강도 산정"
        ws[f'C{base_s8+1}'].value = f"a = As x fy / (0.85 x fck x b) = {ana.a:.3f}mm"
        ws[f'C{base_s8+2}'].value = f"Ø Mn = {label_f} x As x fy x (d - a / 2)"
        ws[f'C{base_s8+3}'].value = f"     = {phi_f_val:.2f} x {ana.as_use:.1f} x {ana.f_y} x ({ana.d_eff:.1f} - {ana.a:.3f} / 2)"
        if ana.M_r > ana.Mu_nm:
            ws[f'C{base_s8+4}'].value = f"     = {ana.M_r:.1f} N.mm ≥ Mu = {ana.Mu_nm:.1f} N.mm  ∴ O.K  [S.F = {ana.M_sf:.3f}]"
        else:
            ws[f'C{base_s8+4}'].value = f"     = {ana.M_r:.1f} N.mm < Mu = {ana.Mu_nm:.1f} N.mm  ∴ N.G  [S.F = {ana.M_sf:.3f}]"

        self._stored_base_s8 = base_s8  # 전단/균열 위치 계산에 활용

    def _write_shear(self, ws, row_offset):
        """10/11) 전단검토"""
        ana = self.analyzer
        spacing_offset = 6 if ana.method == "LSD" else 0
        s11_row = 45 + row_offset + spacing_offset + 5
        if ana.method == "LSD":
            num_shear = 8
        else:
            num_shear = 7
        ws[f'B{s11_row}'].value = f"{num_shear}) 전단검토"
        ws[f'C{s11_row+1}'].value = f"Φ Vc = Φv x \u221afck x b x d / 6"
        ws[f'C{s11_row+2}'].value = f"    = {ana.pi_v:.2f} x \u221a{ana.f_ck} x {ana.beam_b} x {ana.d_eff_v:.1f} / 6 = {ana.pi_V_c:.1f} N"
        if ana.pi_V_c >= ana.Vu_n:
            ws[f'C{s11_row+3}'].value = f"Φ Vc \u2265 Vu = {ana.Vu_n:.1f} N  \u2234 전단보강 불필요"
        else:
            ws[f'C{s11_row+3}'].value = f"Φ Vc < Vu = {ana.Vu_n:.1f} N  \u2234 전단보강 필요"
            ws[f'C{s11_row+5}'].value = f"Av_req = ( {ana.Vu_n/1000:.3f} - {ana.pi_V_c/1000:.3f} ) x {ana.av_space:.1f} / ( {ana.f_y} x {ana.d_eff_v:.1f} x {ana.pi_v:.2f}) = {ana.av_req:.3f} mm\u00b2"
            ws[f'C{s11_row+6}'].value = f"Av_use = {ana.av_use:.3f} mm² ( {ana.rebar_id}{ana.av_dia} - {ana.av_leg}EA,  C.T.C {ana.av_space} mm )"
            if ana.av_space > ana.av_space_min:
                ws[f'C{s11_row+7}'].value = f"사용간격 {ana.av_space} mm > 최소간격 = min(60cm, 0.5D) = {ana.av_space_min:.3f}  ∴ N.G"
            else:
                ws[f'C{s11_row+7}'].value = f"사용간격 {ana.av_space}mm ≤ 최소간격 = min(60cm, 0.5D) = {ana.av_space_min:.3f}  ∴ O.K"
            ws[f'C{s11_row+9}'].value  = f"Vs = {ana.av_use:.3f} x {ana.f_y:.1f} x {ana.d_eff_v:.1f} / {ana.av_space} = {ana.V_s:.1f} N "
            ws[f'C{s11_row+10}'].value = f"Vs_max = 2 x \u221a{ana.f_ck:.1f} / 3 x {ana.beam_b} x {ana.d_eff_v:.3f} "
            if ana.V_s <= ana.V_s_max:
                ws[f'C{s11_row+11}'].value = f"       = {ana.V_s_max:.1f} N ≤ Vs = {ana.V_s:.1f} N  ∴ O.K"
            else:
                ws[f'C{s11_row+11}'].value = f"       = {ana.V_s_max:.1f} N > Vs = {ana.V_s:.1f} N  ∴ N.G"
            if ana.pi_V_n >= ana.Vu_n:
                ws[f'C{s11_row+12}'].value = f" ΦVn = {ana.pi_v:.2f} x ( {ana.V_c:.1f} + {ana.V_s:.1f} ) = {ana.pi_V_n:.3f} N ≥ Vu = {ana.Vu_n:.1f} N  ∴ O.K"
            else:
                ws[f'C{s11_row+12}'].value = f" ΦVn = {ana.pi_v:.2f} x ( {ana.V_c:.1f} + {ana.V_s:.1f} ) = {ana.pi_V_n:.3f} N < Vu = {ana.Vu_n:.1f} N  ∴ N.G"
        self._stored_s11_row = s11_row

    def _write_crack(self, ws, row_offset):
        """11/12) 균열검토"""
        ana = self.analyzer
        crk_row = self._stored_s11_row + 14
        if ana.method == "LSD":
            num_crack = 9
        else:
            num_crack = 8
        ws[f'B{crk_row}'].value = f"{num_crack}) 균열검토"
        ws[f'C{crk_row+1}'].value  = f"① 응력 산정"
        ws[f'D{crk_row+2}'].value  = f"fs = M / [As x (d - χ/3)] = {ana.Ms_nm:.1f} / [ {ana.as_use:.3f} x ( {ana.d_eff:.3f} - {ana.chi_o:.2f} / 3 )] "
        ws[f'D{crk_row+3}'].value  = f"   =  {ana.f_s:.3f} MPa"
        ws[f'D{crk_row+4}'].value  = f"χ = -n x As / b + n x As / b x \u221a [ 1 + 2 x b x d / ( n x As ) ]"
        ws[f'D{crk_row+5}'].value  = f"  = -{ana.nr:.1f} x {ana.as_use:.1f} / {ana.beam_b} + {ana.nr:.1f} x {ana.as_use:.1f} / {ana.beam_b} x \u221a [1 + 2 x {ana.beam_b} x {ana.d_eff:.3f} / ({ana.nr:.1f} x {ana.as_use:.1f})]"
        ws[f'D{crk_row+6}'].value  = f"  = {ana.chi_o:.3f} mm"
        ws[f'D{crk_row+7}'].value  = f"사용철근량 = {ana.as_use:.3f} mm\u00b2  (철근 평균도심 : {ana.d_c:.1f} mm)"
        ws[f'D{crk_row+8}'].value  = f"      1단 : {ana.rebar_id}{ana.as_dia1}-{ana.as_num1:.1f}EA, 2단 : {ana.rebar_id}{ana.as_dia2}-{ana.as_num2:.1f}EA, 3단 : {ana.rebar_id}{ana.as_dia3}-{ana.as_num3:.1f}EA"
        ws[f'C{crk_row+10}'].value = f"② 철근의 최대 중심간격"
        ws[f'D{crk_row+11}'].value = f"강재의 부식에 대한 환경조건은 \u300e {ana.crack_case} \u300f 적용"
        ws[f'D{crk_row+12}'].value = f"Cc = {ana.dc_1:.1f} - {ana.as_dia1} / 2 = {ana.c_c:.2f} mm"
        ws[f'D{crk_row+13}'].value = f"여기서 Cc ; 인장철근이나 긴장재의 표면과 콘크리트 표면사이의 두께(mm)"
        ws[f'D{crk_row+15}'].value = f"Smin : 375 x (Kcr / fs) - 2.5 x Cc = 375 x ({ana.k_cr} / {ana.f_s:.3f}) - 2.5 x {ana.c_c:.3f} = {ana.s_min_1:.3f} mm"
        ws[f'D{crk_row+16}'].value = f"       300 x (Kcr / fs) = 300 x ({ana.k_cr} / {ana.f_s:.3f}) = {ana.s_min_2:.3f} mm"
        ws[f'D{crk_row+17}'].value = f"여기서 Kcr = {ana.k_cr} ( 철근 간격을 통한 균열 검증에서 철근의 노출 조건을 고려한 계수 )"
        ws[f'D{crk_row+18}'].value = f"\u2234 Sa는 작은 값인 {ana.s_min:.3f} mm 를 적용 "
        if ana.s_min >= ana.s_use:
            ws[f'D{crk_row+19}'].value = f" Sa = {ana.s_min:.3f} mm  \u2265 suse = {ana.s_use:.3f} mm  \u2234 O.K"
        else:
            ws[f'D{crk_row+19}'].value = f" Sa = {ana.s_min:.3f} mm  < suse = {ana.s_use:.3f} mm  \u2234 N.G"

    # ── 서브클래스에서 구현 필수 ───────────────────────────────────────────────

    @abstractmethod
    def _write_section6(self, ws) -> int:
        """
        6) 철근비 검토 — 설계법마다 출력이 다르므로 서브클래스에서 구현.
        반환값: row_offset (추가된 행 수, 후속 섹션의 시작 행 계산에 사용됨)
        """
        pass
