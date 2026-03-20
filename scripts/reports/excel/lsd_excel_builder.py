"""
lsd_excel_builder.py
한계상태설계법(LSD) 전용 Excel 보고서 빌더.
"""
from .base_excel_builder import BaseExcelBuilder
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import math

class LSDExcelBuilder(BaseExcelBuilder):
    """한계상태설계법(LSD) 전용 Excel 보고서 빌더."""

    def add_to_workbook(self, wb, sheet_name):
        """LSD 전용 동적 행 구조로 워크북에 시트를 추가한다."""
        wsout = wb.create_sheet(title=sheet_name)
        self._setup_sheet(wsout)
        self._write_header(wsout)
        
        # LSD는 상세 정보가 많아 행 번호를 추적하며 동적으로 작성
        curr = 2 
        curr = self._write_section1_lsd(wsout, curr)
        curr = self._write_section2_lsd(wsout, curr)
        curr = self._write_section3_lsd(wsout, curr)
        curr = self._write_section4_lsd(wsout, curr)
        curr = self._write_section5_lsd(wsout, curr)
        curr = self._write_section6_lsd(wsout, curr)
        curr = self._write_flexure_strength_lsd(wsout, curr)
        curr = self._write_shear_lsd(wsout, curr)
        self._write_crack_lsd_detail(wsout, curr)

    def _merge_row_text(self, ws, row, start_col, end_col, text, align='left', bold=False, color=None):
        """셀을 병합하고 스타일을 적용하여 텍스트를 입력하는 헬퍼 (Section 1 전용)."""
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        cell = ws.cell(row, start_col)
        cell.value = text
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.font = Font(bold=bold, size=9, name='굴림체')
        if color:
            cell.fill = PatternFill(fill_type='solid', fgColor=color)
        return cell

    def _write_row_text(self, ws, row, col, text, bold=False, color=None, align='left'):
        """병합 없이 스타일만 적용하여 텍스트를 입력하는 헬퍼 (Section 2~9 전용)."""
        cell = ws.cell(row, col)
        cell.value = text
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.font = Font(bold=bold, size=9, name='굴림체')
        if color:
            cell.fill = PatternFill(fill_type='solid', fgColor=color)
        return cell

    def _apply_border(self, ws, row, start_col, end_col):
        """병합된 셀 범위에 테두리를 적용."""
        for c in range(start_col, end_col + 1):
            ws.cell(row, c).border = Border(
                left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin')
            )

    def _write_section1_lsd(self, ws, row):
        ana = self.analyzer
        self._merge_row_text(ws, row, 2, 8, '1) 단면제원 및 설계가정', bold=True)
        self._merge_row_text(ws, row+1, 3, 26, f"fck = {ana.f_ck:>5.1f} MPa, fy = {ana.f_y:>5.1f} MPa, \u03a6c = {ana.phi_c:>5.2f}, \u03a6s = {ana.phi_s:>5.2f}, Es = {ana.E_s:>7.0f} MPa")
        
        # Table Header
        headers = [
            ("b(mm)", 3, 5), ("H(mm)", 6, 8), ("d(mm)", 9, 11), ("피복(mm)", 12, 14), 
            ("Mu(kN.m)", 15, 18), ("Vu(kN)", 19, 22), ("Muy(kN.m)", 23, 26)
        ]
        for h, s, e in headers:
            self._merge_row_text(ws, row+2, s, e, h, align='center', bold=True, color='0FFFF0')
            self._apply_border(ws, row+2, s, e)

        # Table Data
        muy_val = ana.Mu
        data = [f"{ana.beam_b:.1f}", f"{ana.beam_h:.1f}", f"{ana.d_eff:.1f}", f"{ana.d_c:.1f}", f"{ana.Mu:.3f}", f"{ana.Vu:.3f}", f"{muy_val:.3f}"]
        for i, val in enumerate(data):
            s, e = headers[i][1], headers[i][2]
            self._merge_row_text(ws, row+3, s, e, val, align='center')
            self._apply_border(ws, row+3, s, e)
        
        return row + 5

    def _write_section2_lsd(self, ws, row):
        ana = self.analyzer
        con = ana.con_material
        self._write_row_text(ws, row, 2, '2) 콘크리트 재료상수', bold=True)
        
        eta_val = ana.alpha_fac / con.alpha_cc
        items = [
            (f"n", f"상승 곡선부의 형상을 나타내는 지수", f"{con.n_eps:>10.2f}"),
            (f"\u03b5co,r", f"최대응력에 처음 도달했을때의 변형률", f"{con.eps_co:>10.4f}"),
            (f"\u03b5cu,r", f"극한변형률", f"{con.eps_cu:>10.4f}"),
            (f"\u03b1cc", f"유효계수", f"{ana.alpha_cc:>10.2f}"),
            (f"fcd", f"콘크리트 설계압축강도 ( fck \u00d7 \u03a6c \u00d7 \u03b1cc )", f"{ana.f_cd:>10.3f} MPa"),
            (f"fcm", f"평균압축강도 ( fck + \u0394f )", f"{con.f_cm:>10.1f} MPa"),
            (f"Ec", f"0.077 mc^1.5 \u221bfcm", f"{con.E_c:>10.1f} MPa"),
            (f"\u03b1", f"압축합력의 평균 응력계수", f"{ana.alpha_fac:>10.2f}"),
            (f"\u03b2", f"압축합력의 작용점 위치계수", f"{ana.beta_fac:>10.2f}"),
            (f"\u03b7", f"등가 사각형 응력 블록의 크기계수", f"{eta_val:>10.2f}"),
            (f"\u03b21", f"등가 사각형 응력 블록의 깊이계수 ( 2\u03b2 )", f"{ana.beta_fac*2:>10.2f}"),
        ]
        for i, (sym, desc, val) in enumerate(items):
            r = row + 1 + i
            self._write_row_text(ws, r, 3, f"  {sym:<8} : {desc:<40} = {val}")
        return row + 1 + len(items) + 1

    def _write_section3_lsd(self, ws, row):
        ana = self.analyzer
        self._write_row_text(ws, row, 2, '3) 철근 재료상수', bold=True)
        self._write_row_text(ws, row+1, 3, f"  fyd    : 설계인장강도 ( \u03a6s fy )                    = {ana.f_yd:>8.1f} MPa")
        self._write_row_text(ws, row+2, 3, f"  \u03b5yd    : 설계 항복 변형률 ( fyd / Es )             = {ana.f_yd/ana.E_s:>8.5f}")
        return row + 4

    def _write_section4_lsd(self, ws, row):
        ana = self.analyzer
        con = ana.con_material
        eta_val = ana.alpha_fac / con.alpha_cc
        self._write_row_text(ws, row, 2, '4) 필요철근량 산정', bold=True)
        self._write_row_text(ws, row+1, 3, "  Mu = As \u00d7 fyd \u00d7 ( d - a/2 )              ---------------- \u2460")
        self._write_row_text(ws, row+2, 3, f"  a = As \u00d7 fyd / ( {eta_val:.2f} \u00d7 fcd \u00d7 b )             ---------------- \u2461")
        self._write_row_text(ws, row+3, 3, "  substitution \u2461 \u2192 \u2460 : solve an equation (As)")
        self._write_row_text(ws, row+4, 3, f"  As_req = {ana.as_req:>8.1f} mm\u00b2")
        return row + 6

    def _write_section5_lsd(self, ws, row):
        ana = self.analyzer
        usage_ratio = ana.as_use / ana.as_req if ana.as_req > 0 else 9.99
        self._write_row_text(ws, row, 2, '5) 사용철근량', bold=True)
        self._write_row_text(ws, row+1, 3, f"* Used As = {ana.as_use:>8.1f} mm\u00b2 (reinforcement center : {ana.d_c:>8.1f} mm), [ 사용율 = {usage_ratio:>6.3f} ]")
        self._write_row_text(ws, row+2, 3, f"  1단 : H {ana.as_dia1:>2} - {ana.as_num1:>3.0f} EA,  Cover = {ana.dc_1:>8.1f} mm, ( {ana.as_use1:>8.1f} mm\u00b2 )")
        if ana.as_num2 > 0:
            self._write_row_text(ws, row+3, 3, f"  2단 : H {ana.as_dia2:>2} - {ana.as_num2:>3.0f} EA,  Cover = {ana.dc_2:>8.1f} mm, ( {ana.as_use2:>8.1f} mm\u00b2 )")
            return row + 5
        return row + 4

    def _write_section6_lsd(self, ws, row):
        ana = self.analyzer
        self._write_row_text(ws, row, 2, '6) 철근비 검토', bold=True)
        ok_min = "O.K" if ana.as_use >= ana.as_min_val else "N.G"
        ok_max = "O.K" if ana.as_use <= ana.as_max_val else "N.G"
        ok_shrink = "O.K" if ana.as_use >= ana.as_shrink else "N.G"

        self._write_row_text(ws, row+1, 3, f"  \u03c1min : 1.4 \u00d7 bw \u00d7 d / fy          = {ana.as_min_1:>8.1f} mm\u00b2")
        self._write_row_text(ws, row+2, 3, f"         0.25 \u221afck \u00d7 bw \u00d7 d / fy  = {ana.as_min_2:>8.1f} mm\u00b2 , As_min = {ana.as_min_val:>8.1f} mm\u00b2")
        self._write_row_text(ws, row+3, 3, f"         4 / 3 \u00d7 Asreq            = {ana.as_min_3:>8.1f} mm\u00b2")
        self._write_row_text(ws, row+4, 3, f"  Shrinkable reinforcemenet       = {ana.as_shrink:>8.1f} mm\u00b2    Shrinkage Reinforcement - {ok_shrink} !!")
        self._write_row_text(ws, row+5, 3, f"  As_max = 0.04 \u00d7 b \u00d7 d           = {ana.as_max_val:>8.1f} mm\u00b2")
        self._write_row_text(ws, row+6, 3, f"  As_max ({ana.as_max_val:.1f}) \u2265 As_use ({ana.as_use:.1f}) \u2265 As_min ({ana.as_min_val:.1f})")
        self._write_row_text(ws, row+7, 3, f"  Max Reinforcement ratio - {ok_max} , Min Reinforcement ratio - {ok_min}")
        self._write_row_text(ws, row+8, 3, f"  철근 간격검토: s_max = min( 2h, 250 ) = {ana.s_detailing_max:.0f} \u2265 s = {ana.s_use:.0f} mm  {ana.s_detailing_ok}")
        return row + 10

    def _write_flexure_strength_lsd(self, ws, row):
        ana = self.analyzer
        con = ana.con_material
        eta_val = ana.alpha_fac / con.alpha_cc
        self._write_row_text(ws, row, 2, "7) 설계 휨강도 산정", bold=True)
        self._write_row_text(ws, row+1, 3, f"  T = As \u00d7 fyd = {ana.as_use:>8.1f} \u00d7 {ana.f_yd:>8.1f} = {ana.tension_force:>10.0f}")
        self._write_row_text(ws, row+2, 3, f"  C = \u03b7 \u00d7 fcd \u00d7 a \u00d7 b = {eta_val:.2f} \u00d7 {ana.f_cd:.3f} \u00d7 a \u00d7 {ana.beam_b:.1f} = {eta_val*ana.f_cd*ana.beam_b:>10.1f} \u00d7 a")
        self._write_row_text(ws, row+3, 3, f"  a = As \u00d7 fyd / ( \u03b7 \u00d7 fcd \u00d7 b ) = {ana.a:>8.3f} mm")
        self._write_row_text(ws, row+4, 3, f"  중립축 깊이 c = a / \u03b21 = {ana.a:>8.1f} / {ana.beta_fac*2:.3f} = {ana.c:>8.1f} mm")
        
        ok_cmax = "O.K !!" if ana.c_max >= ana.c else "N.G !!"
        c_max_calc = f"( {ana.delta_redist:.1f} \u00d7 {ana.eps_cu:.4f} / 0.0033 - 0.600 ) \u00d7 {ana.d_eff:.1f}"
        self._write_row_text(ws, row+5, 3, f"  최대 중립축 깊이 c_max = ( \u03b4 \u00d7 \u03b5cu / 0.0033 - 0.600 ) \u00d7 d")
        self._write_row_text(ws, row+6, 3, f"                         = {c_max_calc}")
        self._write_row_text(ws, row+7, 3, f"                         = {ana.c_max:>8.1f} mm \u2265 c = {ana.c:>8.1f} mm   .. {ok_cmax}")
        
        ok_yield = "O.K !!" if ana.eps_s >= ana.eps_yd else "N.G !!"
        eps_s_calc = f"{ana.eps_cu:.4f} \u00d7 ( {ana.d_eff:.1f} - {ana.c:.1f} ) / {ana.c:.1f}"
        self._write_row_text(ws, row+8, 3, f"  인장철근 변형률 \u03b5s = \u03b5cu \u00d7 ( d - c ) / c")
        self._write_row_text(ws, row+9, 3, f"                   = {eps_s_calc}")
        self._write_row_text(ws, row+10, 3, f"                   = {ana.eps_s:>8.4f} \u2265 \u03b5yd = {ana.eps_yd:>8.4f}   .. {ok_yield}")
        
        res_m = "O.K" if ana.M_r >= ana.Mu_nm else "N.G"
        self._write_row_text(ws, row+11, 3, f"  Mr = {ana.as_use:>8.1f} \u00d7 {ana.f_yd:.0f} \u00d7 ( {ana.d_eff:^7.1f} - a / 2 ) = {ana.M_r/1e6:>8.2f} kN.m")
        self._write_row_text(ws, row+12, 3, f"     \u2265 Mu ( = {ana.Mu:>8.3f} kN.m)  \u2234 {res_m}   [ S.F = {ana.M_sf:>8.3f} ]")
        return row + 14

    def _write_shear_lsd(self, ws, row):
        ana = self.analyzer
        vd = getattr(ana, 'v_details', {})
        self._write_row_text(ws, row, 2, "8) 전단검토 (d = {0:>10.1f} mm)".format(ana.d_eff), bold=True)
        
        vcd_val = vd.get('Vcd_calc', 0)
        vcd_min = vd.get('Vcd_min', 0)
        
        vcd_calc = f"( 0.85 \u00d7 {ana.phi_c:.1f} \u00d7 {vd.get('k',0):.3f} \u00d7 ( {vd.get('rho_l',0):.4f} \u00d7 {ana.f_ck:.0f} )^1/3 + 0.15 \u00d7 {vd.get('fn',0):.2f} )"
        self._write_row_text(ws, row+1, 3, f"  Vcd = ( 0.85 \u00d7 \u03a6c \u00d7 \u03ba \u00d7 (\u03c1 \u00d7 fck)^1/3 + 0.15 \u00d7 fn ) \u00d7 bw \u00d7 d")
        self._write_row_text(ws, row+2, 3, f"      = {vcd_calc}")
        self._write_row_text(ws, row+3, 3, f"        \u00d7 {ana.beam_b:.1f} \u00d7 {ana.d_eff:.1f} = {vcd_val/1e3:>8.1f} kN")
        
        vcd_min_calc = f"( 0.4 \u00d7 {ana.phi_c:.1f} \u00d7 {vd.get('f_ctk',0):.3f} + 0.15 \u00d7 {vd.get('fn',0):.2f} ) \u00d7 {ana.beam_b:.1f} \u00d7 {ana.d_eff:.1f}"
        self._write_row_text(ws, row+4, 3, f"  Vcd,min = ( 0.4 \u00d7 \u03a6c \u00d7 fctk + 0.15 \u00d7 fn ) \u00d7 bw \u00d7 d")
        self._write_row_text(ws, row+5, 3, f"          = {vcd_min_calc}")
        self._write_row_text(ws, row+6, 3, f"          = {vcd_min/1e3:>8.1f} kN")
        
        reinf_needed = " - 전단보강 필요 !!" if ana.pi_V_c < ana.Vu_n else " - 전단보강 불필요"
        self._write_row_text(ws, row+7, 3, f"  \u2234 Vcd = {ana.pi_V_c/1e3:>8.1f} kN  <  Vu = {ana.Vu:>8.1f} kN     : {reinf_needed}")
        
        items = [
            (f"fctk", f"콘크리트 인장강도 (0.7 fctm)", f"{vd.get('f_ctk',0):>8.3f} MPa"),
            (f"\u03ba", f"1 + \u221a( 200 / d ) \u2264 2.0", f"{vd.get('k',0):>8.3f}"),
            (f"\u03c1", f"철근비 As / ( bw \u00d7 d ) \u2264 0.02", f"{vd.get('rho_l',0):>8.4f}"),
            (f"fn", f"Nu / Ac \u2264 0.2 \u03a6c fck", f"{vd.get('fn',0):>8.3f} MPa"),
        ]
        for i, (sym, desc, val) in enumerate(items):
            self._write_row_text(ws, row+8+i, 3, f"     {sym:<6} : {desc:<35} = {val}")
        
        curr = row + 8 + len(items) + 1
        
        if ana.pi_V_c < ana.Vu_n:
            self._write_row_text(ws, curr, 3, "* 전단철근량 검토", bold=True)
            av_use = vd.get('av_use', ana.rebar.get_area(ana.av_dia) * ana.av_leg)
            self._write_row_text(ws, curr+1, 3, f"  Av_use = {av_use:>8.1f} mm\u00b2 ( H {ana.av_dia:>2} - {ana.av_leg:.0f} EA, C.T.C {ana.av_space:.0f} mm)")
            
            vsd_calc = f"{ana.phi_s:.1f} \u00d7 {ana.f_y:.0f} \u00d7 {av_use:.1f} \u00d7 {vd.get('z',0):.1f} \u00d7 {vd.get('cot_theta',0):.3f} / {ana.av_space:.0f}"
            self._write_row_text(ws, curr+2, 3, f"  Vsd = \u03a6s \u00d7 fvy \u00d7 Av \u00d7 z / s \u00d7 cot\u03b8")
            self._write_row_text(ws, curr+3, 3, f"      = {vsd_calc} = {ana.V_s/1e3:>8.1f} kN")
            
            tan_theta = 1/vd.get('cot_theta',1) if vd.get('cot_theta',0) > 0 else 0
            vd_max_calc = f"{vd.get('nu',0):.3f} \u00d7 {ana.phi_c:.1f} \u00d7 {ana.f_ck:.0f} \u00d7 {ana.beam_b:.1f} \u00d7 {vd.get('z',0):.1f} / {vd.get('cot_theta',0)+tan_theta:.3f}"
            vd_max_val = vd.get('Vdmax2',0)
            self._write_row_text(ws, curr+4, 3, f"  Vd,Max = {vd_max_calc}")
            self._write_row_text(ws, curr+5, 3, f"         = {vd_max_val/1e3:>8.1f} kN \u2265 Vsd = {ana.V_s/1e3:>8.1f} kN .. {'O.K' if vd_max_val >= ana.V_s else 'N.G'}")
            
            rho_v = av_use / (ana.av_space * ana.beam_b * math.sin(math.radians(vd.get('alpha_deg',90)))) if (ana.av_space * ana.beam_b) > 0 else 0
            rho_v_min = (0.08 * math.sqrt(ana.f_ck)) / ana.f_y if ana.f_y > 0 else 0
            self._write_row_text(ws, curr+6, 3, f"  \u03c1v_use = {rho_v:>8.5f} \u2265 \u03c1v_min = {rho_v_min:>8.5f} .. {'O.K' if rho_v >= rho_v_min else 'N.G'}")
            
            s_ok = "O.K" if ana.av_space <= vd.get('s_max_1',0) else "N.G"
            self._write_row_text(ws, curr+7, 3, f"  Space Check : s = {ana.av_space:>8.0f} mm \u2264 s_max = {vd.get('s_max_1',0):>8.0f} mm .. {s_ok}")
            
            v_total = (ana.pi_V_c + ana.V_s) / 1e3
            v_ok = "O.K" if v_total >= ana.Vu else "N.G"
            self._write_row_text(ws, curr+8, 3, f"  Vd = {v_total:>8.1f} kN \u2265 {ana.Vu:.1f} kN .. {v_ok} [ S.F = {v_total/ana.Vu if ana.Vu>0 else 9.99:.3f} ]")
            
            self._write_row_text(ws, curr+10, 3, "* 종방향 철근의 추가인장력 검토", bold=True)
            dt_calc = f"0.5 \u00d7 {ana.Vu:>8.3f} \u00d7 ( {vd.get('cot_theta',0):.3f} - {vd.get('cot_alpha',0):.3f} )"
            self._write_row_text(ws, curr+11, 3, f"  \u0394T = 0.5 \u00d7 Vu \u00d7 ( cot\u03b8 - cot\u03b1 ) = {dt_calc}")
            self._write_row_text(ws, curr+12, 3, f"     = {ana.delta_t/1e3:>8.3f} kN")
            
            dtb_calc = f"( {ana.M_r/1e6:.3f} - {ana.Mu_nm/1e6:.3f} ) / {vd.get('z',0)/1000:^7.3f}"
            self._write_row_text(ws, curr+13, 3, f"  \u0394TB = ( Mr - Mu ) / z = {dtb_calc}")
            self._write_row_text(ws, curr+14, 3, f"      = {ana.delta_tb/1e3:>8.3f} kN")
            
            t_ok = "O.K" if ana.delta_t <= ana.delta_tb else "N.G"
            self._write_row_text(ws, curr+15, 3, f"  \u2234 \u0394T { '\u2264' if t_ok=='O.K' else '>' } \u0394TB .. {t_ok}")
            return curr + 17
        else:
            return curr + 1

    def _write_crack_lsd_detail(self, ws, row):
        ana = self.analyzer
        sd = getattr(ana, 'service_details', {})
        if not sd: return
        self._write_row_text(ws, row, 2, "9) 균열 및 철근 간격 검토", bold=True)
        
        # 1. Min Rebar
        self._write_row_text(ws, row+1, 3, "\u2460 최소철근량 검토", bold=True)
        as_min_calc = f"{sd.get('kc',0.4):.2f} \u00d7 {sd.get('k_scale',1.0):.2f} \u00d7 {sd.get('Act',0):.0f} \u00d7 {sd.get('fctm',0):.2f} / {ana.f_y:.0f}"
        self._write_row_text(ws, row+2, 4, f"As_min = kc \u00d7 k \u00d7 Act \u00d7 fct / fs")
        self._write_row_text(ws, row+3, 4, f"       = {as_min_calc} = {sd.get('as_min_lsd',0):>8.2f} mm\u00b2")
        status_min = "O.K" if ana.as_use >= sd.get('as_min_lsd', 0) else "N.G"
        self._write_row_text(ws, row+4, 4, f"As_use = {ana.as_use:>8.2f} mm\u00b2 \u2265 As_min .. {status_min}")
        
        # 2. Indirect Crack Control
        self._write_row_text(ws, row+6, 3, "\u2461 간접균열제어 (Stress fs)", bold=True)
        fs_calc = f"{sd.get('Ms_knm',0)*1e3:.3f} / ( {ana.as_use:.1f} \u00d7 ( {ana.d_eff:.1f} - {sd.get('c_neutral',0):.1f} / 3 ) )"
        self._write_row_text(ws, row+7, 4, f"fs = Ms / ( As \u00d7 ( d - c / 3 ) )")
        self._write_row_text(ws, row+8, 4, f"   = {fs_calc} = {sd.get('fs',0):>8.3f} MPa")
        self._write_row_text(ws, row+9, 4, f"fs \u2264 fsa = 0.8 \u00d7 fy = {sd.get('fsa',0):.1f} MPa .. {'O.K' if sd.get('fs',0) <= sd.get('fsa',0) else 'N.G'}")
        
        # 3. Max Bar Diameter
        self._write_row_text(ws, row+11, 3, "\u2462 철근직경검토", bold=True)
        dia_limit = sd.get('max_dia_limit', 0)
        dia_use = ana.as_dia1
        self._write_row_text(ws, row+12, 4, f"\u03a6_limit = {dia_limit:.1f} mm (fs = {sd.get('fs',0):.1f} MPa 기준)")
        self._write_row_text(ws, row+13, 4, f"\u03a6_use = {dia_use:.0f} mm \u2264 \u03a6_limit = {dia_limit:.1f} mm .. {'O.K' if dia_use <= dia_limit else 'N.G'}")
        
        # 4. Spacing
        self._write_row_text(ws, row+15, 3, "\u2463 철근 간격검토", bold=True)
        sa_val = sd.get('sa_limit', 300.0)
        s_use = ana.beam_b / ana.as_num1 if ana.as_num1 > 0 else 0
        self._write_row_text(ws, row+16, 4, f"S = {ana.beam_b:.0f} / {ana.as_num1:.1f} EA = {s_use:.1f} mm \u2264 Sa = {sa_val:.1f} mm .. {'O.K' if s_use <= sa_val else 'N.G'}")

    # Base overrides (unused in LSD but required to avoid mess)
    def _write_section6(self, ws): return 0
    def _write_flexure_strength(self, ws, offset): pass
    def _write_shear(self, ws, offset): pass
    def _write_crack(self, ws, offset): pass
