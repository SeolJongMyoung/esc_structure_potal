from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from string import ascii_uppercase

class ExcelReportBuilder:
    """
    RCSectionAnalyzer의 연산 결과를 받아 엑셀 보고서를 생성하는 클래스.
    """
    def __init__(self, analyzer):
        self.analyzer = analyzer
        self.std = analyzer.standard

    def add_to_workbook(self, wb, sheet_name):
        wsout = wb.create_sheet(title=sheet_name)
        
        # Set up basic formatting
        alpalist = list(ascii_uppercase)
        for i in range(1, 100):
            wsout.row_dimensions[i].height = 15
        for i in alpalist:
            wsout.column_dimensions[i].width = 3.0
        font_format = Font(size=9, name='굴림체')
        for rows in wsout["A1":"Z100"]:
            for cell in rows:
                cell.font = font_format

        # Header Title
        wsout.merge_cells('B1:M1')
        wsout['B1'].value = f"[{self.std.name}] - RC 단면 검토 보고서"
        wsout['B1'].font = Font(size=11, bold=True, name='굴림체')
        wsout['B1'].alignment = Alignment(horizontal='left')

        # Section 1: Basic Information
        wsout['B2'].value = '1) 단면제원 및 설계가정'
        label_f = "Øc" if self.analyzer.method == "LSD" else "Øf"
        label_v = "Øs" if self.analyzer.method == "LSD" else "Øv"
        phi_f_val = self.analyzer.phi_c if self.analyzer.method == "LSD" else self.analyzer.pi_f
        phi_v_val = self.analyzer.phi_s if self.analyzer.method == "LSD" else self.analyzer.pi_v
        
        wsout['C3'].value = f"fck = {self.analyzer.f_ck} MPa, fy = {self.analyzer.f_y} MPa, {label_f} = {phi_f_val:.2f}, {label_v} = {phi_v_val:.2f}, Es = {self.analyzer.E_s} MPa"

        # Formatting for table headers
        for r in range(4,6) :                        
            for c in range(3,24) :                      
                wsout.cell(r,c).alignment = Alignment(horizontal='center', vertical='center')  
        for r in range(4,6) :                        
            for c in range(3,24) :                      
                wsout.cell(r,c).border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        for r in range(4,6) :                        
            for c in [3,6,9,12] :                       
                c1 = c + 2
                wsout.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1)
        for r in range(4,6) :                         
            for c in [15,19,23] :                          
                c1 = c + 3
                wsout.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1)    
        for c in range(3,24) :                         
            wsout.cell(4,c).fill = PatternFill(fill_type='solid', fgColor='0FFFF0')
        
        wsout['C4'].value = 'B(mm)'; wsout['F4'].value = 'H(mm)'; wsout['I4'].value = 'd(mm)'; wsout['L4'].value = '피복(mm)'
        wsout['O4'].value = 'Mu(N.mm)'; wsout['S4'].value = 'Vu(N)'; wsout['W4'].value = 'Ms(N.mm)'
        wsout['C5'].value = self.analyzer.beam_b; wsout['F5'].value = self.analyzer.beam_h; wsout['I5'].value = f"{self.analyzer.d_eff:.1f}"; wsout['L5'].value = f"{self.analyzer.d_c:.1f}"
        wsout['O5'].value = self.analyzer.Mu_nm; wsout['S5'].value = self.analyzer.Vu_n; wsout['W5'].value = self.analyzer.Ms_nm

        wsout['B7'].value = '2) 콘크리트 재료상수'
        wsout['C8'].value = 'β1    : 등가 사각형 응력 블록의 깊이계수'; wsout['P8'].value = '='; wsout['Q8'].value = self.analyzer.beta_1

        # Section 3: Strength reduction factor
        wsout['B10'].value = '3) 강도감소계수(Ø) 산정'
        wsout['C11'].value = f"T = As x fy = {self.analyzer.as_use:.3f} x {self.analyzer.f_yd:.1f} = {self.analyzer.tension_force:.1f} N" if self.analyzer.method == "LSD" else f"T = As x fy = {self.analyzer.as_use:.3f} x {self.analyzer.f_y:.3f} = {self.analyzer.tension_force:.1f} N"
        
        if self.analyzer.method == "LSD":
            wsout['C12'].value = f"C = alpha_cc x phi_c x fck x alpha x b = {self.analyzer.alpha_cc:.2f} x {self.analyzer.phi_c:.2f} x {self.analyzer.f_ck} x {self.analyzer.alpha_fac:.2f} x {self.analyzer.beam_b} = {self.analyzer.compression_force:.1f} x c"
        else:
            wsout['C12'].value = f"C = 0.85 x fck x a x b = 0.85 x {self.analyzer.f_ck:.3f} x a x {self.analyzer.beam_b:.3f} = {self.analyzer.compression_force:.1f} x a"
        wsout['C13'].value = f"T = C 이므로, a = {self.analyzer.a:.3f} mm, c = {self.analyzer.a:.3f} / β1 = {self.analyzer.a:.3f} / {self.analyzer.beta_1:.3f} = {self.analyzer.c:.3f} mm"
        wsout['C14'].value = f"εy = fy / Es = {self.analyzer.f_y} / {self.analyzer.E_s} = {self.analyzer.epsilon_y:.5f}"
        wsout['C15'].value = f"εt = 0.00300 x (dt - c) / c = 0.00300 x ({self.analyzer.d_eff:.3f} - {self.analyzer.c:.2f}) / {self.analyzer.c:.2f} = {self.analyzer.epsilon_t:.5f}"
        if self.analyzer.epsilon_t >= 0.005:
            wsout['C16'].value = f"εt ≥ 0.0050 이므로 {self.analyzer.epsilon_t_result}이며, Ø = {self.analyzer.pi_f_r:.2f} 를 적용한다"
        else:
            wsout['C16'].value = f"εt < 0.0050 이므로 {self.analyzer.epsilon_t_result}이며, Ø = {self.analyzer.pi_f_r:.2f} 를 적용한다"

        # Section 4: Required Reinforcement Calculation
        wsout['B18'].value = '4) 필요철근량 산정'
        wsout['C19'].value = 'Mu / Øf = As x fy  x (d - a / 2)              ----------------   ①'
        wsout['C20'].value = ' a = As x fy  / ( 0.85 x fck x b)             ----------------   ②'
        wsout['C21'].value = ' 식②를 식①에 대입하여 이차방정식으로 As를 구한다'
        wsout['E22'].value = ' fy²                                Mu'
        wsout['C23'].value = f" ────────── As² - fy x d x As + ───  = 0 ,   Asreq = {self.analyzer.as_req:.3f} mm²"
        wsout['C24'].value = ' 2 x 0.85 x fck x b                         Øf '
        
        # Section 5: Used Reinforcement
        wsout['B26'].value = f"5) 사용철근량 : Asuse = {self.analyzer.as_use:.1f} mm², 철근도심 : dc_avg = {self.analyzer.d_c:.1f}mm,  [ 사용율 = {self.analyzer.as_use/self.analyzer.as_req if self.analyzer.as_req > 0 else 0:.3f} ]"
        wsout['F27'].value = f"1단 : {self.analyzer.rebar_id} {self.analyzer.as_dia1} - {self.analyzer.as_num1} EA (= {self.analyzer.as_use1:.1f} mm², dc1 = {self.analyzer.dc_1:.1f} mm)"
        wsout['F28'].value = f"2단 : {self.analyzer.rebar_id} {self.analyzer.as_dia2} - {self.analyzer.as_num2} EA (= {self.analyzer.as_use2:.1f} mm², dc2 = {self.analyzer.dc_2:.1f} mm)"
        wsout['F29'].value = f"3단 : {self.analyzer.rebar_id} {self.analyzer.as_dia3} - {self.analyzer.as_num3} EA (= {self.analyzer.as_use3:.1f} mm², dc3 = {self.analyzer.dc_3:.1f} mm)"

        # Section 6: Reinforcement Check
        wsout['B31'].value = "6) 철근비 검토"
        
        row_offset = 0
        if ("콘크리트" in self.std.name or "KCI" in self.std.name) and hasattr(self.analyzer, 'min_rebar_res') and 'mcr' in self.analyzer.min_rebar_res:
             mcr_knm = self.analyzer.min_rebar_res.get('mcr', 0) / 1e6
             limit_knm = self.analyzer.min_rebar_res.get('limit', 0) / 1e6
             ig = self.analyzer.min_rebar_res.get('ig', 0)
             fr = self.analyzer.min_rebar_res.get('fr', 0)
             phi_mn_knm = self.analyzer.M_r / 1e6
             
             wsout['C32'].value = "* 최소 철근량 검토"
             wsout['C33'].value = f"Ig = bh³ / 12 = {self.analyzer.beam_b:.2f} x {self.analyzer.beam_h:.2f}³ / 12 = {ig:,.0f} mm⁴"
             wsout['C34'].value = f"fr = 0.63λ\u221a(fck) = 0.63 x 1.0 x \u221a({self.analyzer.f_ck}) = {fr:.3f}"
             wsout['C35'].value = f"Mcr = fr x Ig / yt = {fr:.3f} x {ig:,.0f} / ({self.analyzer.beam_h:.2f} / 2) = {mcr_knm:,.3f} kN.m"
             
             compare_sign = "<" if 1.2 * mcr_knm < (4/3) * (self.analyzer.Mu_nm/1e6) else "\u2265"
             wsout['C36'].value = f"1.2Mcr = {1.2*mcr_knm:,.3f} kN.m {compare_sign} 4/3Mu = {(4/3)*(self.analyzer.Mu_nm/1e6):,.3f} kN.m"
             
             status_min = "만족" if self.analyzer.min_rebar_res['is_ok'] else "불만족"
             ok_ng_min = "O.K" if self.analyzer.min_rebar_res['is_ok'] else "N.G"
             wsout['C37'].value = f"\u03a6Mn = {phi_mn_knm:,.3f} kN.m \u2265 {limit_knm:,.3f} kN.m  최소철근량 {status_min}, \u2234 {ok_ng_min}"
             
             wsout['C39'].value = "* 최대 철근비 검토"
             rho_max = self.analyzer.max_rebar_res['rho_max']
             rho_use = self.analyzer.max_rebar_res['rho_use']
             status_max = "만족" if self.analyzer.max_rebar_res['is_ok'] else "불만족"
             ok_ng_max = "O.K" if self.analyzer.max_rebar_res['is_ok'] else "N.G"
             wsout['C40'].value = f"\u03c1max = {rho_max:.5f}"
             wsout['C41'].value = f"\u03c1use = {rho_use:.5f} (As / bd)"
             compare_sign_max = "<" if rho_use < rho_max else "\u2265"
             wsout['C42'].value = f"\u03c1use {compare_sign_max} \u03c1max \u2192 철근비 {status_max}, \u2234 {ok_ng_max}"
             row_offset = 6
        else:
            wsout['C32'].value = f"ρmin : 1.4 / fy          = {self.analyzer.lo_min_1:.6f} "
            wsout['C33'].value = f"       0.25 x √fck / fy  = {self.analyzer.lo_min_2:.6f},  ρmin = {self.analyzer.lo_min:.6f} 적용"
            wsout['C34'].value = f"ρmax = 0.75 x ρb = 0.75 x (0.85 x β1 x fck / fy) x (6,000 / (6,000 + fy)) = {self.analyzer.lo_bal:.6f}" 
            wsout['C35'].value = f"ρuse = As / ( b x d ) = {self.analyzer.lo_use:.6f} "
            if self.analyzer.lo_use >= self.analyzer.lo_min :
                if self.analyzer.lo_use < self.analyzer.lo_max :
                    wsout['C36'].value = f"ρmax ≥ ρuse ≥ ρmin --> 최소철근비, 최대철근비 만족   ∴ O.K"
                else:
                    wsout['C36'].value = f"ρmax < ρuse ≥ ρmin --> 최소철근비 만족, 최대 철근비 불만족   ∴ N.G"
            else :
                if self.analyzer.lo_use < self.analyzer.lo_max :
                    wsout['C36'].value = f"ρmax ≥ ρuse < ρmin --> 최소철근비 불만족,  최대 철근비 만족   ∴ N.G"
                    if self.analyzer.lo_use >= self.analyzer.lo_min_3 :
                        wsout['C37'].value = f"ρuse ≥ 4 x ρreq / 3 = {self.analyzer.lo_min_3:.6f} --> 최소철근비 만족   ∴ O.K"
                    else :
                        wsout['C37'].value = f"ρuse < 4 x ρreq / 3 = {self.analyzer.lo_min_3:.6f} --> 최소철근비 불만족   ∴ N.G"
                else :    
                    wsout['C36'].value = f"ρmax < ρuse < ρmin --> 최소철근비, 최대 철근비 불만족   ∴ N.G"

        # Section 7: Spacing Check
        base_s7 = 39 + row_offset
        wsout[f'B{base_s7}'].value = "7) 철근 간격 검토"
        wsout[f'C{base_s7+1}'].value = f"s_max = min( 2h, 250 ) = {self.analyzer.s_detailing_max:.0f} mm"
        wsout[f'C{base_s7+2}'].value = f"s = b / n = {self.analyzer.beam_b:.1f} / {self.analyzer.as_num1} = {self.analyzer.s_use:.1f} mm"
        compare_sign_s = "\u2265" if self.analyzer.s_detailing_max >= self.analyzer.s_use else "<"
        wsout[f'C{base_s7+3}'].value = f"s_max {compare_sign_s} s  \u2234 {self.analyzer.s_detailing_ok}"
        
        spacing_offset = 6 # Added rows for Section 7
        
        # Section 8: Design Flexural Strength Calculation
        base_s8 = 39 + row_offset + spacing_offset
        wsout[f'B{base_s8}'].value = "8) 설계 휨강도 산정"
        wsout[f'C{base_s8+1}'].value = f"a = As x fy / (0.85 x fck x b) = {self.analyzer.a:.3f}mm"
        wsout[f'C{base_s8+2}'].value = f"Ø Mn = {label_f} x As x fy x (d - a / 2)"
        wsout[f'C{base_s8+3}'].value = f"     = {phi_f_val:.2f} x {self.analyzer.as_use:.1f} x {self.analyzer.f_y} x ({self.analyzer.d_eff:.1f} - {self.analyzer.a:.3f} / 2)"
        if self.analyzer.M_r > self.analyzer.Mu_nm:
            wsout[f'C{base_s8+4}'].value = f"     = {self.analyzer.M_r:.1f} N.mm ≥ Mu = {self.analyzer.Mu_nm:.1f} N.mm  ∴ O.K  [S.F = {self.analyzer.M_sf:.3f}]"
        else:
            wsout[f'C{base_s8+4}'].value = f"     = {self.analyzer.M_r:.1f} N.mm < Mu = {self.analyzer.Mu_nm:.1f} N.mm  ∴ N.G  [S.F = {self.analyzer.M_sf:.3f}]"

        # Section 11: Shear Check (was 10)
        s11_row = 45 + row_offset + spacing_offset + 5 # Offset for strength check rows
        wsout[f'B{s11_row}'].value ="11) 전단검토"
        wsout[f'C{s11_row+1}'].value =f"Φ Vc = Φv x \u221afck x b x d / 6"
        wsout[f'C{s11_row+2}'].value =f"    = {self.analyzer.pi_v:.2f} x \u221a{self.analyzer.f_ck} x {self.analyzer.beam_b} x {self.analyzer.d_eff_v:.1f} / 6 = {self.analyzer.pi_V_c:.1f} N"
        if self.analyzer.pi_V_c >= self.analyzer.Vu_n :
            wsout[f'C{s11_row+3}'].value =f"Φ Vc \u2265 Vu = {self.analyzer.Vu_n:.1f} N  \u2234 \uc804\ub2e8\ubcf4\uac15 \ubd88\ud544\uc694"
        else :
            wsout[f'C{s11_row+3}'].value =f"Φ Vc < Vu = {self.analyzer.Vu_n:.1f} N  \u2234 \uc804\ub2e8\ubcf4\uac15 \ud544\uc694"
            wsout[f'C{s11_row+5}'].value =f"Av_req = ( {self.analyzer.Vu_n/1000:.3f} - {self.analyzer.pi_V_c/1000:.3f} ) x {self.analyzer.av_space:.1f} / ( {self.analyzer.f_y} x {self.analyzer.d_eff_v:.1f} x {self.analyzer.pi_v:.2f}) = {self.analyzer.av_req:.3f} mm\u00b2"
            wsout[f'C{s11_row+6}'].value =f"Av_use = {self.analyzer.av_use:.3f} mm² ( {self.analyzer.rebar_id}{self.analyzer.av_dia} - {self.analyzer.av_leg}EA,  C.T.C {self.analyzer.av_space} mm )"
            if self.analyzer.av_space > self.analyzer.av_space_min :
                wsout[f'C{s11_row+7}'].value =f"사용간격 {self.analyzer.av_space} mm > 최소간격 = min(60cm, 0.5D) = {self.analyzer.av_space_min:.3f}  ∴ N.G"
            else :
                wsout[f'C{s11_row+7}'].value =f"사용간격 {self.analyzer.av_space}mm ≤ 최소간격 = min(60cm, 0.5D) = {self.analyzer.av_space_min:.3f}  ∴ O.K"
            wsout[f'C{s11_row+9}'].value =f"Vs = {self.analyzer.av_use:.3f} x {self.analyzer.f_y:.1f} x {self.analyzer.d_eff_v:.1f} / {self.analyzer.av_space} = {self.analyzer.V_s:.1f} N "
            wsout[f'C{s11_row+10}'].value =f"Vs_max = 2 x √{self.analyzer.f_ck:.1f} / 3 x {self.analyzer.beam_b} x {self.analyzer.d_eff_v:.3f} "

            if self.analyzer.V_s <= self.analyzer.V_s_max :
                wsout[f'C{s11_row+11}'].value =f"       = {self.analyzer.V_s_max:.1f} N ≤ Vs = {self.analyzer.V_s:.1f} N  ∴ O.K" 
            else :
                wsout[f'C{s11_row+11}'].value =f"       = {self.analyzer.V_s_max:.1f} N > Vs = {self.analyzer.V_s:.1f} N  ∴ N.G" 

            if self.analyzer.pi_V_n >= self.analyzer.Vu_n :
                wsout[f'C{s11_row+12}'].value =f" ΦVn = {self.analyzer.pi_v:.2f} x ( {self.analyzer.V_c:.1f} + {self.analyzer.V_s:.1f} ) = {self.analyzer.pi_V_n:.3f} N ≥ Vu = {self.analyzer.Vu_n:.1f} N  ∴ O.K" 
            else :
                wsout[f'C{s11_row+12}'].value =f" ΦVn = {self.analyzer.pi_v:.2f} x ( {self.analyzer.V_c:.1f} + {self.analyzer.V_s:.1f} ) = {self.analyzer.pi_V_n:.3f} N < Vu = {self.analyzer.Vu_n:.1f} N  ∴ N.G"
            
        # Section 12: Crack Control (was 11)
        crk_row = s11_row + 14
        wsout[f'B{crk_row}'].value ="12) 균열검토"
        wsout[f'C{crk_row+1}'].value =f"① 응력 산정"
        wsout[f'D{crk_row+2}'].value =f"fs = M / [As x (d - \u03c7/3)] = {self.analyzer.Ms_nm:.1f} / [ {self.analyzer.as_use:.3f} x ( {self.analyzer.d_eff:.3f} - {self.analyzer.chi_o:.2f} / 3 )] "
        wsout[f'D{crk_row+3}'].value =f"   =  {self.analyzer.f_s:.3f} MPa"
        wsout[f'D{crk_row+4}'].value =f"\u03c7 = -n x As / b + n x As / b x \u221a [ 1 + 2 x b x d / ( n x As ) ]"
        wsout[f'D{crk_row+5}'].value =f"  = -{self.analyzer.nr:.1f} x {self.analyzer.as_use:.1f} / {self.analyzer.beam_b} + {self.analyzer.nr:.1f} x {self.analyzer.as_use:.1f} / {self.analyzer.beam_b} x \u221a [1 + 2 x {self.analyzer.beam_b} x {self.analyzer.d_eff:.3f} / ({self.analyzer.nr:.1f} x {self.analyzer.as_use:.1f})]"
        wsout[f'D{crk_row+6}'].value =f"  = {self.analyzer.chi_o:.3f} mm"
        wsout[f'D{crk_row+7}'].value = f"사용철근량 = {self.analyzer.as_use:.3f} mm\u00b2  (철근군 평균도심 : {self.analyzer.d_c:.1f} mm)"
        wsout[f'D{crk_row+8}'].value = f"      1단 : {self.analyzer.rebar_id}{self.analyzer.as_dia1}-{self.analyzer.as_num1:.1f}EA, 2단 : {self.analyzer.rebar_id}{self.analyzer.as_dia2}-{self.analyzer.as_num2:.1f}EA, 3단 : {self.analyzer.rebar_id}{self.analyzer.as_dia3}-{self.analyzer.as_num3:.1f}EA"
        
        wsout[f'C{crk_row+10}'].value =f"\u2461 \ucc20\uadf9\uc758 \ucd5c\ub300 \uc911\uc2ec\uac04\uaca9"
        wsout[f'D{crk_row+11}'].value =f"\uac15\uc7ac\uc758 \ubd80\uc2dd\uc5d0 \ub300\ud55c \ud658\uacbd\uc870\uac74\uc740 \u300e {self.analyzer.crack_case} \u300f \uc801\uc6a9"
        wsout[f'D{crk_row+12}'].value =f"Cc = {self.analyzer.dc_1:.1f} - {self.analyzer.as_dia1} / 2 = {self.analyzer.c_c:.2f} mm"
        wsout[f'D{crk_row+13}'].value =f"\uc5ec\uae30\uc11c Cc ; \uc778\uc7a5\ucc20\uadf9\uc774\ub098 \uae34\uc7a5\uc7ac\uc758 \ud45c\uba74\uacfc \ucabd\ud06c\ub9ac\ud2b8 \ud45c\uba74\uc0ac\uc774\uc758 \ud450\uaed8(mm)"
        wsout[f'D{crk_row+15}'].value =f"Smin : 375 x (Kcr / fs) - 2.5 x Cc = 375 x ({self.analyzer.k_cr} / {self.analyzer.f_s:.3f}) - 2.5 x {self.analyzer.c_c:.3f} = {self.analyzer.s_min_1:.3f} mm"
        wsout[f'D{crk_row+16}'].value =f"       300 x (Kcr / fs) = 300 x ({self.analyzer.k_cr} / {self.analyzer.f_s:.3f}) = {self.analyzer.s_min_2:.3f} mm"
        wsout[f'D{crk_row+17}'].value =f"\uc5ec\uae30\uc11c Kcr = {self.analyzer.k_cr} ( \ucc20\uadf9 \uac04\uaca9\uc744 \ud1b5\ud55c \uade0\uc5f4 \uac80\uc99d\uc5d0\uc11c \ucc20\uadf9\uc758 \ub178\ucd9c \uc870\uac74\uc744 \uace0\ub824\ud55c \uacc4\uc218 )" 
        wsout[f'D{crk_row+18}'].value =f"\u2234 Sa\ub294 \uc791\uc740 \uac12\uc778 {self.analyzer.s_min:.3f} mm \ub9bc \uc801\uc6a9 "
        
        if self.analyzer.s_min >= self.analyzer.s_use :
            wsout[f'D{crk_row+19}'].value =f" Sa = {self.analyzer.s_min:.3f} mm  \u2265 suse = {self.analyzer.s_use:.3f} mm  \u2234 O.K" 
        else:
            wsout[f'D{crk_row+19}'].value =f" Sa = {self.analyzer.s_min:.3f} mm  < suse = {self.analyzer.s_use:.3f} mm  \u2234 N.G"
